# FM POP extrage stocurile din fisierele cu Kalculation ruckgabe
# Se seteaza ca active sheet primul sheet
from openpyxl import load_workbook

print("="*34,"\n= PROGRAM CARE EXTRAGE STOCURILE =\n= DIN FISELE DE MAGAZIE POP      =","\n"+"="*34, "\n")
cat_ini = input("Introduceti o categorie:  ")
cat_fin = cat_ini+".xlsx"
wb = load_workbook(cat_fin, data_only = True) # data_only = ce se vede in celula; nu formula
wss = wb.get_sheet_names() # preia sheets from wb
#print(wss) # afiseaza toate sheet-urile
dictionar = {}
for i in wss:
	ws = wb.get_sheet_by_name(i) # intram in ws

	ultimul_rand = ws.max_row
	cell_name = "J{}".format(ultimul_rand)
	if ws[cell_name].value == 0:
		pass
	else:
		a1 = str(ws).strip('<')
		a2 = a1.strip('>')
		a3 = a2[10:].strip('"') # pana aici curata (<Workbook ">); ramane OK
		valoare = ws[cell_name].value
		print("{} = {}".format(a3, valoare)) #afiseaza sheet si valoare
		dictionar[a3] = valoare
print(dictionar)
