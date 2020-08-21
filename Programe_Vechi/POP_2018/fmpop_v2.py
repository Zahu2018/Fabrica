# FM POP extrage stocurile din fisierele cu Kalculation ruckgabe
# Se seteaza ca active sheet primul sheet
from openpyxl import load_workbook
import tkinter
import tkinter.filedialog
from tkinter import messagebox
import os
class App():
    def __init__(self, master):
        self.master = master
        b = tkinter.Button(master, text = 'Deschide fisierul', command=self.dc)
        b.grid(row=0, column=0)
        self.l = tkinter.Label(master, text = 'CATEGORIE MATERIALE')
        self.l.grid(row=1, column=0)
    #========================================
    def dc(self): #actiunea butonului OPEN -> butoane categorii
        self.filename = tkinter.filedialog.askopenfilename()
        cat_fin = self.filename
        head, tail = os.path.split(self.filename) # head=fila.ext; tail = path
        self.l.config(text=tail)
        fila = open("rezultat.csv", "a")
        try:
            wb = load_workbook(cat_fin, data_only = True) # data_only = ce se vede in celula; nu formula
            wss = wb.get_sheet_names() # preia sheets from wb
            #print(wss) # afiseaza toate sheet-urile
            dictionar = {}
            for i in wss:
                ws = wb.get_sheet_by_name(i) # intram in ws
            
                ultimul_rand = ws.max_row
                cell_name = "J{}".format(ultimul_rand)
                if ws[cell_name].value == 0:
                    pass
                else:
                    a1 = str(ws).strip('<')
                    a2 = a1.strip('>')
                    a3 = a2[10:].strip('"') # pana aici curata (<Workbook ">); ramane OK
                    valoare = ws[cell_name].value
                    rez = ("{} = {}".format(a3, valoare)) #afiseaza sheet si valoare
                    print(rez)
                    dictionar[a3] = valoare
            #print(dictionar)
        except IndexError:
            messagebox.showinfo("ATENTIE" , "Redeschide si salveaza fisierul inca o data")
        except FileNotFoundError:
            messagebox.showinfo("ATENTIE" , "Numele fisierului nu este scris corect.\nFii atent la ce scrii.")
        except PermissionError:
            messagebox.showinfo("ATENTIE" , "Inchide fisierul {} si mai porneste programul odata.\nVezi sa nu fie deschis fisierul cu o alta aplicatie.".format(cat_ini))
                    
        fila = open("rezultat.csv", "a")
        r = 2
        for key, value in sorted(dictionar.items()):
            fi = str(key)+','+str(value/100)+'\n'
            tkinter.Label(master, text=fi).grid(row=r, column=0)
            r += 1
            fila.write(fi)
        fila.close()
        
root = tkinter.Tk()
app = App(root)
root.mainloop()
