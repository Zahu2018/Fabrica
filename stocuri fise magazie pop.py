"""EXTRAGE STOCURI FISE MAGAZIE POP."""

from openpyxl import load_workbook
import os

"""{'Author': 'Florian Zah',
'Contact': '',
'Copyright': '',
'Credits': '',
'Date': '24.05.2019',
'Description': 'Extrage stocurile din fisierele cu Kalculation ruckgabe Pop',
'Last Modification': '27.09.2019',
'Licence': '',
'Maintainer': 'Florian Zah',
'Status': 'Production',
'Tags': ['stocuri', 'pop', 'ruckgabe', 'extrage', 'fise magazie'],
'Title': 'EXTRAGE STOCURI FM POP',
'Version': '1.0.0'}.
"""

# Se seteaza ca active sheet primul sheet

fisiere1 = os.listdir()
#  Cautare doar in fisierele python: .xlsx
fisiere = []
for fi in fisiere1:
    if fi[-5:] == '.xlsx':
        fii = fi.replace('.xlsx', '')
        fisiere.append(fii)

print("=" * 34, "\n= PROGRAM CARE EXTRAGE STOCURILE =\n= DIN FISELE DE MAGAZIE POP      =", "\n" + "=" * 34, "\n")
print(fisiere, '\n')
cat_ini = input("Introduceti o categorie:  ")
cat_fin = cat_ini + ".xlsx"
fila = open("rezultat.csv", "a")
try:
    wb = load_workbook(cat_fin, data_only=True)  # data_only = ce se vede in celula; nu formula
    wss = wb.sheetnames  # preia sheets from wb
    # print(wss)  # afiseaza toate sheet-urile
    dictionar = {}
    for i in wss:
        ws = wb[i]  # intram in ws; wb.get_sheet_by_name(i) = vechi

        ultimul_rand = ws.max_row
        cell_name = "J{}".format(ultimul_rand)
        if ws[cell_name].value == 0:
            pass
        else:
            a1 = str(ws).strip('<')
            a2 = a1.strip('>')
            a3 = a2[10:].strip('"')  # pana aici curata (<Workbook ">); ramane OK
            valoare = ws[cell_name].value
            rez = ("{} = {}".format(a3, valoare))  # afiseaza sheet si valoare
            print(rez)
            dictionar[a3] = valoare
    # print(dictionar)
except IndexError:
    print(1 * '\n')
    print(50 * '=')
    print("Redeschide si salveaza fisierul inca odata.")
    print(50 * '=')
    print(2 * '\n')
except FileNotFoundError:
    print(1 * '\n')
    print(50 * '=')
    print("Numele fisierului nu este scris corect.\nFii atent la ce scrii.")
    print(50 * '=')
    print(2 * '\n')
except PermissionError:
    print(1 * '\n')
    print(50 * '=')
    print("Inchide fisierul {} si mai porneste programul odata.\nVezi sa nu fie deschis fisierul cu o alta aplicatie.".format(cat_ini))
    print(50 * '=')
    print(2 * '\n')

fila = open("rezultat.csv", "a")
for key, value in sorted(dictionar.items()):
    fi = str(key) + ',' + str(value / 100) + '\n'
    fila.write(fi)
fila.close()
