"""ABLIEFERUNG."""
# NU MODIFICA ABSOLUT NIMIC !!!

from openpyxl import load_workbook
import tkinter
from tkinter import messagebox, ttk
import os


"""{'Author': 'Florian Zah',
'Contact': '',
'Copyright': '',
'Credits': '',
'Date': '07.08.2019',
'Deprecated': '',
'Description': 'Face o copia a fisierului ABLIEFERUNG VALENTI pentru raport cu noile date',
'Last Modification': '07.10.2019',
'Licence': '',
'Maintainer': 'Florian Zah',
'Status': 'Production',
'Tags': ['ablieferung', 'raport', 'creeaza', 'gui'],
'Title': 'ABLIEFERUNG',
'Version': '2.1.0'}.
"""

ABLIEFERUNG_STAS = "ABLIEFERUNG VALENTI - .xlsx"


def deschide_xls():
    """Citeste Sheet Valenti."""
    data = a.get()
    wb = load_workbook(ABLIEFERUNG_STAS, data_only=False)  # False = nu pierzi formula
    # wss = wb.sheetnames  #  preia sheets from wb
    # print(wss)  #  afiseaza toate sheet-urile
    ws = wb.active
    ws['E2'] = data
    # ws.cell(row=2, column=5).value = 2
    # ws.cell(coordinate="E2").value = 3  # 'coordinate=' is optional here

    # salveaza xlsx-ul cu noua data si noul nume
    nume_fisier = "C:\\Users\\User\\Desktop\\ABLIEFERUNG VALENTI - " + data + ".xlsx"
    ablieferung_uri = os.listdir("C:\\Users\\User\\Desktop\\")
    ablieferung_nou = "ABLIEFERUNG VALENTI - " + data + ".xlsx"
    if ablieferung_nou in ablieferung_uri:
            messagebox.showinfo("INFORMARE", 'Ablieferung cu data de\\nn{:^30}\n\nexista deja.\nSchimbati data.'.format(data))
    else:
        wb.save(nume_fisier)
        messagebox.showinfo("INFORMARE", 'A fost creat fisierul\nABLIEFERUNG cu data de\n\n{:^30}'.format(data))
        a.delete(0, 'end')


def interfata():
    """Creaza interfata GUI."""
    global a, b
    app = tkinter.Tk()
    app.title('Ablieferung generator')
    lab = tkinter.Label(app, text='Introduceti data:')
    lab.grid(row=0, column=0)
    a = ttk.Entry(app)
    a.grid(column=0, row=1)
    b = ttk.Button(app, text='Creaza Ablieferung', command=deschide_xls)
    b.grid(column=1, row=1)
    app.mainloop()


if __name__ == "__main__":
    interfata()
