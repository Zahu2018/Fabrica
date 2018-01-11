'''
#citeste o coloana din excel
from openpyxl import load_workbook
wb = load_workbook('Piele.xlsx')
wss = wb.get_sheet_names() # preia sheets from wb
ws = wb.get_sheet_by_name('NUBUK TAUPE')
for row in range(1, ws.max_row):
	for column in "D":
		cell_name = "{}{}".format(column, row)
		a = ws[cell_name].value
		print(a)
'''
'''
#citeste o matrice (parte) din tabel
from openpyxl import load_workbook
wb = load_workbook('Piele.xlsx')
wss = wb.get_sheet_names() # preia sheets from wb
ws = wb.get_sheet_by_name('NUBUK TAUPE')
l1 = [] #lista cu toate randurile
for row in range(4, ws.max_row):
	l = [] #lista pe un rand
	for column in "DEFGHI":
		cell_name = "{}{}".format(column, row)
		a = ws[cell_name].value
		if a == None:
			pass
		else:
			l.append(a)
	
	l1.append(l)
print(l1)
'''
# FM POP extrage stocurile din fisierele cu Kalculation ruckgabe
from openpyxl import load_workbook
print("="*34,"\n= PROGRAM CARE EXTRAGE STOCURILE =\n= DIN FISELE DE MAGAZIE POP      =","\n"+"="*34, "\n")
unu  = input("  1. Extrage stocurile")
doi  = input("  2. Pregateste pt. ruckgabe")
trei = input("  3. Scade stocurile")
cat_ini = input("Introduceti o categorie:  ")
cat_fin = cat_ini+".xlsx"
wb = load_workbook(cat_fin, data_only = True) # data_only = ce se vede in celula; nu formula
wss = wb.get_sheet_names() # preia sheets from wb
print(wss)
for i in wss:
	ws = wb.get_sheet_by_name(i) # intram in ws
	ultimul_rand = ws.max_row
	cell_name = "J{}".format(ultimul_rand)
	if ws[cell_name].value == 0:
		pass
	else:
		valoare = ws[cell_name].value
		print("{} = {}".format(ws, valoare))
