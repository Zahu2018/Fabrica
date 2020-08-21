#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""{'Author': 'Florian Zah',
 'Contact': '',
 'Copyright': '',
 'Credits': '',
 'Date': '24.05.2019',
 'Deprecated': '',
 'Description': 'Face situatie pt o ata sau toate',
 'Last Modification': '',
 'Licence': '',
 'Maintainer': '',
 'Status': '',
 'Tags': ['serafil', 'ata', 'situatie', 'fineplanung', 'bind', 'event'],
 'Title': 'SITUATIE SERAFIL',
 'Version': '1'}
."""

import tkinter
from tkinter import N, S, E, W
from openpyxl import load_workbook

path = "D:\\Valenti\\2019\\Feinplanung.xlsx"
dicti = {}

def citeste_ods():
    '''Citeste saison4 si saison5 intr-o lista mare'''
    wb = load_workbook(path, data_only = True) # data_only = ce se vede in celula; nu formula
    #wss = wb.sheetnames # preia sheets from wb
    #print(wss) # afiseaza toate sheet-urile
    lista_toala = []
    for i in range(2): # activam sheet pe rand: 0 = primul sheet; 1 = al doilea sheet; etc
        wb.active = i # activa primul sheet
        ws = wb.active # sheet din care citeste
        m_row=ws.max_row # get max row count
        m_column=ws.max_column # get max column count
        lista_mare = []
        for i in range(2,m_row+1): # iterate over all cells and rows
            lista_linie = [] # lista pt fiecare linie
            for j in range(1,m_column+1): # iterate over all columns
                cell_obj=ws.cell(row=i,column=j) # get particular cell value
                lista_linie.append(cell_obj.value)
            lista_mare.append(lista_linie)
        lista_toala.extend(lista_mare) # extinde lista NU APPEND
            #print(lista_linie)
            #print(cell_obj.value,end=',', '\n') # print cell value
    #print(lista_toala)
    fa_matricea(lista_toala)

def fa_matricea(lista):
    '''Face o matrice dintr-o lista de liste - din Feinplanung'''
    global L
    L = [] # lista de matrice
    M = [] # matrice
    for item in lista:
        if str(item[9]).isupper():
            if M != []:
                L.append(M)
            else:
                pass
            M = []
            M.append(item)
        else:
            M.append(item)
    M.append(item)
    L.append(M)
    #print(L)
    #label_text.set('OK')
    Fa_dictionar(L)



def Fa_dictionar(L):

    for mat in L:
        for item in mat[0]:
            if "SERAFIL" in str(item):
                ind = mat[0].index(item) # gasim indexul fiecarui item cu "SERAFIL"
                #print(item, ' = ', ind)
                l = len(mat) # lungimea (inaltimea) matricei in linii
                for i in range(2 , l):
                    suma = 0
                    suma_i=0
                    m = mat[i][ind] #
                    try:
                        sm = float(m)
                        suma += sm
                        suma_i += int(suma)
                        if item in dicti:
                            var = int(dicti.get(item))  # a fost float???
                            suma_i += var
                            dicti[item] = suma_i
                        else:
                            dicti[item] = suma_i
                    except:
                        print('nu pot converti in float', item)
                        label_avertisment.set(item)
    #print(dicti)

def Serafil_unu(dicti):
    '''Face pt. o singura ata'''
    at = entr.get() # ata pt care vrem raport
    ata = 'SERAFIL '+ str(at)
    if ata in dicti.keys():
        #print(ata, dicti.get(ata))
        label_text.set('{} = {} m'.format(ata,dicti.get(ata)))
    else:
        #print("Aceasta ata nu exista pentru productie")
        label_text.set('Aceasta ata nu este in productie\nIncearca cu 0 sau 00 inainte de numar serafil')
    entr.delete(0, 'end')

def Serafil_tot(dicti):
    '''Face pt. toate atele'''
    listare = ''
    listare_1 = ''
    for item in dicti:
        lis = ''
        lis = (str(item) + "," + str(dicti.get(item)))
        lis_1 = (str(item) + " = " + str(dicti.get(item))) # pentru gui
        #print(item, " = ", (dicti.get(item)))
        listare += str(lis) + '\n'
        listare_1 += str(lis_1) + '\n'
    #print(listare)
    label_text_rap.set(listare_1)

    var=open("Situatie_Serafil.csv", 'w') #numerele trebuie convertite in siruri inainte de scriere in fisier [str()]
    var.write(listare)
    var.close()
    print("Raportul cu Serafil a fost creat.")


def gui():
    global avertisment_label, rez, rap, entr, un_ser, sit_ser, label_text, label_text_rap, label_avertisment
    master = tkinter.Tk()
    master.title('Calculator necesar ata')
    master.geometry("350x100")
    f1 = tkinter.Frame(master)
    f1.grid(row=0, column=0, ipadx=1)
    f2 = tkinter.Frame(master)
    f2.grid(row=1, column=0)


    entr = tkinter.Entry(f1)
    entr.grid(row=0, column=0, sticky=W)
    entr.focus()

    un_ser = tkinter.Button(f1, text='Calculeaza', command=lambda : Serafil_unu(dicti))
    un_ser.grid(row=0, column=1, sticky=E, padx=2)
    master.bind("<Return>", lambda event: Serafil_unu(dicti))
    sit_ser = tkinter.Button(f1, text='Genereaza raport', command=lambda : Serafil_tot(dicti))
    sit_ser.grid(row=0, column=2, sticky=E, padx=2)

    label_avertisment = tkinter.StringVar()
    label_text = tkinter.StringVar()
    label_text_rap = tkinter.StringVar()

    avertisment_label = tkinter.Label(f2, textvariable=label_avertisment)
    avertisment_label.grid(row=0, column=0)
    rez = tkinter.Label(f2, textvariable=label_text, font=('arial', 12, 'bold'))
    rez.grid(row=1, column=0, sticky=N+S+E+W)
    rap = tkinter.Label(f2, textvariable=label_text_rap)
    rap.grid(row=2, column=0, sticky=N+S+E+W)
    #label_text.set("Asteapta !")

    citeste_ods()
    master.mainloop()

gui()
