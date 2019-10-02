"""ABLIEFERUNG."""

from openpyxl import load_workbook

"""{'Author': 'Florian Zah',
'Contact': '',
'Copyright': '',
'Credits': '',
'Date': '07.08.2019',
'Deprecated': '',
'Description': 'Face o copia a fisierului ABLIEFERUNG VALENTI pentru raport cu noile date',
'Last Modification': '27.09.2019',
'Licence': '',
'Maintainer': 'Florian Zah',
'Status': 'Production',
'Tags': ['ablieferung', 'raport', 'creeaza'],
'Title': 'ABLIEFERUNG',
'Version': '1.1.0'}.
"""


ABLIEFERUNG_STAS = "ABLIEFERUNG VALENTI - .xlsx"


def creeaza_ablieferung():
    """Preia data."""
    data = input("Scrie data pentru care vrei sa creezi\n"
                 "Ablieferung si apasa tasta Enter:\n")
    deschide_xls(data)


def deschide_xls(data):
    """Citeste Sheet Valenti."""
    wb = load_workbook(ABLIEFERUNG_STAS, data_only=False)  # False = nu pierzi formula
    # wss = wb.sheetnames  #  preia sheets from wb
    # print(wss)  #  afiseaza toate sheet-urile
    ws = wb.active
    ws['E2'] = data
    # ws.cell(row=2, column=5).value = 2
    # ws.cell(coordinate="E2").value = 3  # 'coordinate=' is optional here

    # salveaza xlsx-ul cu noua data si noul nume
    nume_fisier = "ABLIEFERUNG VALENTI - " + data + '.xlsx'
    wb.save(nume_fisier)
    meniu(data)


def meniu(data):
    """Meniu."""
    print("\nFisierul ABLIEFERUNG VALENTI - {} a fost creat.".format(data))
    print("Apasa:\n1 si apoi ENTER - Pentru a crea un nou Ablieferung; sau\n"
          "0 si apoi ENTER - Pentru a inchide programul\n")
    x = input()
    if x == '1':
        creeaza_ablieferung()
    elif x == '0':
        exit(0)
    else:
        print("Ai apasat tasta '{}'. Te rog sa apesi 0 sau 1".format(x))
        meniu(data=0)


if __name__ == "__main__":
    creeaza_ablieferung()
